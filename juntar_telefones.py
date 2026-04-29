import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font, Alignment

# 👇 Muda aqui para o caminho do seu arquivo
arquivo = 'C:/Users/User/Documents/OneDrive/Desktop/WA Sender/telefones.xlsx'

# Lê a planilha
df = pd.read_excel(arquivo, dtype=str)

# Junta todos os telefones em uma string separada por vírgula
todos_juntos = ', '.join(df['telefone'].tolist())

# Abre o arquivo para editar
wb = load_workbook(arquivo)
ws = wb.active

# Coloca tudo na célula C1
ws['C1'] = todos_juntos
ws['C1'].font = Font(name='Arial')
ws['C1'].alignment = Alignment(wrap_text=True)
ws.column_dimensions['C'].width = 60

# Salva (troca o nome para não sobrescrever o original!)
wb.save('resultado.xlsx')

print("Feito! Abra o arquivo resultado.xlsx 🎉")